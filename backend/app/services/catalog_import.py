import asyncio
import csv
import io
import re
import zipfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions.catalog import (
    CatalogImportFileError,
    CatalogImportTooLargeError,
    CatalogImportValidationError,
)
from app.schemas.catalog import (
    CatalogImportFileMetadata,
    CatalogImportPreviewResponse,
    CatalogImportPreviewRow,
    CatalogImportRowError,
    CatalogItemCreate,
)
from app.services.catalog import find_existing_catalog_skus

MAX_CATALOG_IMPORT_BYTES: Final = 10 * 1024 * 1024
MAX_CATALOG_IMPORT_ROWS: Final = 2_000
CATALOG_IMPORT_PREVIEW_ROWS: Final = 100
MAX_XLSX_UNCOMPRESSED_BYTES: Final = 100 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES: Final = 10_000

_HEADER_ALIASES: Final[dict[str, frozenset[str]]] = {
    "name": frozenset(
        {"name", "product name", "product", "service name", "title", "item"}
    ),
    "item_type": frozenset({"type", "item type", "product type"}),
    "description": frozenset({"description", "details", "product description"}),
    "sku": frozenset({"sku", "code", "product code", "item code"}),
    "price": frozenset({"price", "unit price", "sale price"}),
    "status": frozenset({"status"}),
}
_WHITESPACE_PATTERN: Final = re.compile(r"\s+")


@dataclass(frozen=True, slots=True)
class _RawImportRow:
    row_number: int
    values: dict[str, Any]


@dataclass(frozen=True, slots=True)
class _ParsedImportFile:
    metadata: CatalogImportFileMetadata
    detected_columns: dict[str, str]
    rows: list[_RawImportRow]


@dataclass(slots=True)
class PreparedCatalogImportRow:
    row_number: int
    normalized: dict[str, Any]
    item: CatalogItemCreate | None
    errors: list[CatalogImportRowError]


@dataclass(frozen=True, slots=True)
class PreparedCatalogImport:
    preview: CatalogImportPreviewResponse
    items: list[CatalogItemCreate]
    rows: list[PreparedCatalogImportRow]


async def prepare_catalog_import(
    session: AsyncSession,
    business_id: UUID,
    *,
    filename: str | None,
    content: bytes,
) -> PreparedCatalogImport:
    """Parse, normalize, and validate an upload without writing any rows."""
    parsed = await asyncio.to_thread(parse_catalog_import_file, filename, content)
    prepared_rows = [_validate_row(row) for row in parsed.rows]
    _add_upload_duplicate_errors(prepared_rows)

    candidate_skus = {
        sku for row in prepared_rows if (sku := _candidate_sku(row)) is not None
    }
    existing_skus = await find_existing_catalog_skus(
        session,
        business_id,
        candidate_skus,
    )
    _add_existing_sku_errors(prepared_rows, existing_skus)

    all_errors = sorted(
        (error for row in prepared_rows for error in row.errors),
        key=lambda error: (error.row, error.field or "", error.message),
    )
    invalid_row_numbers = {error.row for error in all_errors}
    valid_items = [
        row.item
        for row in prepared_rows
        if row.item is not None and row.row_number not in invalid_row_numbers
    ]
    preview = CatalogImportPreviewResponse(
        file=parsed.metadata,
        detected_columns=parsed.detected_columns,
        total_rows=len(prepared_rows),
        valid_rows=len(prepared_rows) - len(invalid_row_numbers),
        invalid_rows=len(invalid_row_numbers),
        preview_rows=[
            CatalogImportPreviewRow(
                row=row.row_number,
                normalized=row.normalized,
                item=(
                    row.item
                    if row.item is not None
                    and row.row_number not in invalid_row_numbers
                    else None
                ),
                errors=row.errors,
            )
            for row in prepared_rows[:CATALOG_IMPORT_PREVIEW_ROWS]
        ],
        errors=all_errors,
        preview_limit=CATALOG_IMPORT_PREVIEW_ROWS,
    )
    return PreparedCatalogImport(
        preview=preview,
        items=[item for item in valid_items if item is not None],
        rows=prepared_rows,
    )


def require_valid_catalog_import(
    prepared: PreparedCatalogImport,
) -> list[CatalogItemCreate]:
    """Return normalized creates or raise the typed atomic-validation error."""
    if prepared.preview.invalid_rows:
        raise CatalogImportValidationError(prepared.preview)
    return prepared.items


def parse_catalog_import_file(
    filename: str | None,
    content: bytes,
) -> _ParsedImportFile:
    safe_filename = Path(filename or "").name
    if not safe_filename:
        raise CatalogImportFileError("The upload must have a filename")
    if len(content) > MAX_CATALOG_IMPORT_BYTES:
        raise CatalogImportTooLargeError(
            "Catalog import files must be 10 MB or smaller"
        )
    if not content:
        raise CatalogImportFileError("The catalog import file is empty")

    suffix = Path(safe_filename).suffix.lower()
    if suffix == ".csv":
        headers, rows = _parse_csv(content)
        file_type = "csv"
    elif suffix == ".xlsx":
        headers, rows = _parse_xlsx(content)
        file_type = "xlsx"
    else:
        raise CatalogImportFileError("Only CSV and XLSX files are supported")

    detected_columns, column_indexes = _detect_columns(headers)
    if "name" not in detected_columns:
        raise CatalogImportFileError("The file must include a recognizable name column")
    if not rows:
        raise CatalogImportFileError("The file does not contain any catalog rows")
    if len(rows) > MAX_CATALOG_IMPORT_ROWS:
        raise CatalogImportFileError(
            f"Catalog imports may contain at most {MAX_CATALOG_IMPORT_ROWS} rows"
        )

    canonical_rows = [
        _RawImportRow(
            row_number=row_number,
            values={
                canonical: values[index] if index < len(values) else None
                for canonical, index in column_indexes.items()
            },
        )
        for row_number, values in rows
    ]
    return _ParsedImportFile(
        metadata=CatalogImportFileMetadata(
            filename=safe_filename[:255],
            file_type=file_type,
            size_bytes=len(content),
        ),
        detected_columns=detected_columns,
        rows=canonical_rows,
    )


def _parse_csv(content: bytes) -> tuple[list[Any], list[tuple[int, list[Any]]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise CatalogImportFileError("CSV files must use UTF-8 encoding") from None

    try:
        reader = csv.reader(io.StringIO(text, newline=""), strict=True)
        header: list[Any] | None = None
        rows: list[tuple[int, list[Any]]] = []
        for values in reader:
            if header is None:
                if _row_is_blank(values):
                    continue
                header = values
                continue
            if _row_is_blank(values):
                continue
            rows.append((reader.line_num, values))
            if len(rows) > MAX_CATALOG_IMPORT_ROWS:
                break
    except (csv.Error, UnicodeError):
        raise CatalogImportFileError("The CSV file is malformed") from None

    if header is None:
        raise CatalogImportFileError("The CSV file has no usable header row")
    return header, rows


def _parse_xlsx(content: bytes) -> tuple[list[Any], list[tuple[int, list[Any]]]]:
    _validate_xlsx_archive(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (InvalidFileException, OSError, ValueError, KeyError, zipfile.BadZipFile):
        raise CatalogImportFileError("The XLSX file is invalid") from None
    except (AttributeError, EOFError, RuntimeError, TypeError):
        raise CatalogImportFileError("The XLSX file is invalid") from None

    try:
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            header: list[Any] | None = None
            rows: list[tuple[int, list[Any]]] = []
            for row_number, cells in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                values = list(cells)
                if header is None:
                    if _row_is_blank(values):
                        continue
                    header = values
                    continue
                if _row_is_blank(values):
                    continue
                rows.append((row_number, values))
                if len(rows) > MAX_CATALOG_IMPORT_ROWS:
                    break
            if header is not None and rows and _header_has_name(header):
                return header, rows
    except (OSError, ValueError, KeyError, zipfile.BadZipFile):
        raise CatalogImportFileError("The XLSX file is invalid") from None
    except (AttributeError, EOFError, RuntimeError, TypeError):
        raise CatalogImportFileError("The XLSX file is invalid") from None
    finally:
        workbook.close()

    raise CatalogImportFileError("The XLSX file has no usable visible worksheet")


def _validate_xlsx_archive(content: bytes) -> None:
    stream = io.BytesIO(content)
    if not zipfile.is_zipfile(stream):
        raise CatalogImportFileError("The XLSX file is invalid")
    try:
        with zipfile.ZipFile(stream) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise CatalogImportFileError("The XLSX file is too complex")
            if sum(entry.file_size for entry in entries) > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise CatalogImportFileError("The XLSX file is too large when expanded")
            if any(entry.flag_bits & 0x1 for entry in entries):
                raise CatalogImportFileError("Encrypted XLSX files are not supported")
            if any("vbaproject" in entry.filename.lower() for entry in entries):
                raise CatalogImportFileError(
                    "Macro-enabled workbooks are not supported"
                )
    except zipfile.BadZipFile:
        raise CatalogImportFileError("The XLSX file is invalid") from None


def _detect_columns(headers: list[Any]) -> tuple[dict[str, str], dict[str, int]]:
    detected: dict[str, str] = {}
    indexes: dict[str, int] = {}
    for index, raw_header in enumerate(headers):
        original = "" if raw_header is None else str(raw_header).strip()
        normalized = _normalize_header(original)
        if not normalized:
            continue
        canonical = next(
            (
                field
                for field, aliases in _HEADER_ALIASES.items()
                if normalized in aliases
            ),
            None,
        )
        if canonical is None:
            continue
        if canonical in detected:
            raise CatalogImportFileError(
                f"Multiple columns map to the supported field '{canonical}'"
            )
        detected[canonical] = original
        indexes[canonical] = index
    return detected, indexes


def _header_has_name(headers: list[Any]) -> bool:
    return any(
        _normalize_header("" if header is None else str(header))
        in _HEADER_ALIASES["name"]
        for header in headers
    )


def _validate_row(row: _RawImportRow) -> PreparedCatalogImportRow:
    normalized = {
        "name": _normalize_cell(row.values.get("name")),
        "item_type": _normalize_cell(row.values.get("item_type")) or "product",
        "description": _normalize_optional_cell(row.values.get("description")),
        "sku": _normalize_optional_cell(row.values.get("sku")),
        "price": _normalize_optional_cell(row.values.get("price")),
        "status": _normalize_cell(row.values.get("status")) or "active",
    }
    normalized["item_type"] = normalized["item_type"].lower()
    normalized["status"] = normalized["status"].lower()
    if normalized["sku"] is not None:
        normalized["sku"] = normalized["sku"].upper()
    try:
        item = CatalogItemCreate.model_validate(normalized)
    except ValidationError as error:
        errors = [
            _render_validation_error(row.row_number, issue) for issue in error.errors()
        ]
        return PreparedCatalogImportRow(
            row_number=row.row_number,
            normalized=normalized,
            item=None,
            errors=errors,
        )
    return PreparedCatalogImportRow(
        row_number=row.row_number,
        normalized=normalized,
        item=item,
        errors=[],
    )


def _add_upload_duplicate_errors(rows: list[PreparedCatalogImportRow]) -> None:
    counts = Counter(sku for row in rows if (sku := _candidate_sku(row)) is not None)
    duplicates = {sku for sku, count in counts.items() if count > 1}
    for row in rows:
        if _candidate_sku(row) in duplicates:
            row.errors.append(
                CatalogImportRowError(
                    row=row.row_number,
                    field="sku",
                    message="sku is duplicated within the uploaded file",
                )
            )


def _add_existing_sku_errors(
    rows: list[PreparedCatalogImportRow],
    existing_skus: set[str],
) -> None:
    for row in rows:
        if _candidate_sku(row) in existing_skus:
            row.errors.append(
                CatalogImportRowError(
                    row=row.row_number,
                    field="sku",
                    message="sku already exists in this business",
                )
            )


def _candidate_sku(row: PreparedCatalogImportRow) -> str | None:
    sku = row.normalized.get("sku")
    if not isinstance(sku, str) or not sku or len(sku) > 100:
        return None
    return sku


def _render_validation_error(
    row_number: int,
    issue: dict[str, Any],
) -> CatalogImportRowError:
    location = issue.get("loc", ())
    field = str(location[0]) if location else None
    issue_type = issue.get("type")
    message = str(issue.get("msg", "value is invalid")).lower()
    if issue_type == "missing":
        message = f"{field or 'value'} is required"
    elif issue_type == "string_too_short" and field == "name":
        message = "name is required"
    elif issue_type == "greater_than_equal" and field == "price":
        message = "price must be greater than or equal to 0"
    elif issue_type in {"decimal_max_places", "decimal_whole_digits"}:
        message = "price must fit NUMERIC(14, 2)"
    elif field is not None and not message.startswith(field):
        message = f"{field} {message}"
    return CatalogImportRowError(
        row=row_number,
        field=field,
        message=message,
    )


def _normalize_header(value: str) -> str:
    return _WHITESPACE_PATTERN.sub(" ", value.strip().lower())


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_optional_cell(value: Any) -> str | None:
    normalized = _normalize_cell(value)
    return normalized or None


def _row_is_blank(values: list[Any]) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in values
    )
