import io
import os
import unittest
from decimal import Decimal
from unittest.mock import patch
from uuid import UUID

from openpyxl import Workbook

os.environ.setdefault(
    "AIBOS_DATABASE_URL",
    "postgresql+asyncpg://database.invalid/test",
)
os.environ.setdefault("AIBOS_AUTH_SECRET_KEY", "x" * 32)

from app.exceptions.catalog import (
    CatalogImportFileError,
    CatalogImportTooLargeError,
)
from app.services.catalog_import import (
    CATALOG_IMPORT_PREVIEW_ROWS,
    MAX_CATALOG_IMPORT_BYTES,
    parse_catalog_import_file,
    prepare_catalog_import,
)

BUSINESS_A = UUID("00000000-0000-0000-0000-00000000000a")
BUSINESS_B = UUID("00000000-0000-0000-0000-00000000000b")


class CatalogImportPreviewTests(unittest.IsolatedAsyncioTestCase):
    async def test_csv_preview_detects_aliases_defaults_and_normalizes(self) -> None:
        content = (
            b"Product Name,Details,Product Code,Sale Price\n"
            b"  Widget  ,  Useful  , abc-1 ,19.95\n"
        )
        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.csv",
            content=content,
        )

        preview = prepared.preview
        self.assertEqual(
            preview.detected_columns,
            {
                "name": "Product Name",
                "description": "Details",
                "sku": "Product Code",
                "price": "Sale Price",
            },
        )
        self.assertEqual(
            (preview.total_rows, preview.valid_rows, preview.invalid_rows), (1, 1, 0)
        )
        item = preview.preview_rows[0].item
        self.assertIsNotNone(item)
        assert item is not None
        self.assertEqual(item.name, "Widget")
        self.assertEqual(item.item_type, "product")
        self.assertEqual(item.status, "active")
        self.assertEqual(item.sku, "ABC-1")
        self.assertEqual(item.price, Decimal("19.95"))
        self.assertEqual(preview.preview_rows[0].normalized["sku"], "ABC-1")

    async def test_xlsx_preview_uses_first_visible_usable_sheet(self) -> None:
        workbook = Workbook()
        hidden = workbook.active
        hidden.title = "Hidden"
        hidden.sheet_state = "hidden"
        sheet = workbook.create_sheet("Catalog")
        sheet.append(["Service Name", "Type", "Unit Price", "Status"])
        sheet.append(["Consultation", "service", 125, "draft"])

        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.xlsx",
            content=_workbook_bytes(workbook),
        )

        item = prepared.items[0]
        self.assertEqual(item.name, "Consultation")
        self.assertEqual(item.item_type, "service")
        self.assertEqual(item.status, "draft")
        self.assertEqual(item.price, Decimal(125))
        self.assertEqual(prepared.preview.file.file_type, "xlsx")

    async def test_required_name_is_enforced_with_human_row_number(self) -> None:
        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.csv",
            content=b"name,price\n,10.00\n",
        )

        self.assertEqual(prepared.preview.invalid_rows, 1)
        self.assertEqual(prepared.preview.errors[0].row, 2)
        self.assertEqual(prepared.preview.errors[0].field, "name")
        self.assertEqual(prepared.preview.errors[0].message, "name is required")

    async def test_negative_price_is_rejected_with_friendly_error(self) -> None:
        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.csv",
            content=b"name,price\nWidget,-0.01\n",
        )
        self.assertEqual(
            prepared.preview.errors[0].message,
            "price must be greater than or equal to 0",
        )
        self.assertEqual(prepared.preview.errors[0].row, 2)

    async def test_invalid_status_and_item_type_are_rejected(self) -> None:
        content = b"name,item type,status\nWidget,inventory,deleted\n"
        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.csv",
            content=content,
        )
        fields = {error.field for error in prepared.preview.errors}
        self.assertEqual(fields, {"item_type", "status"})
        self.assertEqual(prepared.preview.invalid_rows, 1)

    async def test_duplicate_sku_inside_upload_marks_every_conflicting_row(
        self,
    ) -> None:
        content = b"name,sku\nOne,dupe\nTwo, DUPE \n"
        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.csv",
            content=content,
        )
        self.assertEqual(prepared.preview.invalid_rows, 2)
        self.assertEqual([error.row for error in prepared.preview.errors], [2, 3])
        self.assertTrue(
            all("uploaded file" in error.message for error in prepared.preview.errors)
        )

    async def test_sku_conflicts_are_reported_even_with_other_row_errors(
        self,
    ) -> None:
        prepared = await prepare_catalog_import(
            _SkuSession({BUSINESS_A: {"DUPE"}}),
            BUSINESS_A,
            filename="catalog.csv",
            content=b"name,sku\n,dupe\nValid,dupe\n",
        )
        errors_by_row = {
            row: [
                error.message for error in prepared.preview.errors if error.row == row
            ]
            for row in (2, 3)
        }
        self.assertIn("name is required", errors_by_row[2])
        for row in (2, 3):
            self.assertIn(
                "sku is duplicated within the uploaded file",
                errors_by_row[row],
            )
            self.assertIn("sku already exists in this business", errors_by_row[row])

    async def test_existing_sku_conflict_is_scoped_to_business(self) -> None:
        session = _SkuSession({BUSINESS_A: {"EXISTING"}, BUSINESS_B: {"OTHER"}})
        prepared = await prepare_catalog_import(
            session,
            BUSINESS_A,
            filename="catalog.csv",
            content=b"name,sku\nOne,existing\nTwo,other\n",
        )
        self.assertEqual(prepared.preview.invalid_rows, 1)
        self.assertEqual(prepared.preview.valid_rows, 1)
        self.assertEqual(prepared.preview.errors[0].row, 2)
        self.assertEqual(prepared.items[0].sku, "OTHER")
        self.assertEqual(session.requested_business_id, BUSINESS_A)

    async def test_preview_is_bounded_but_counts_all_rows_and_errors(self) -> None:
        rows = ["name,price"] + [f"Item {index},-1" for index in range(101)]
        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.csv",
            content=("\n".join(rows) + "\n").encode(),
        )
        self.assertEqual(prepared.preview.total_rows, 101)
        self.assertEqual(prepared.preview.invalid_rows, 101)
        self.assertEqual(len(prepared.preview.errors), 101)
        self.assertEqual(
            len(prepared.preview.preview_rows), CATALOG_IMPORT_PREVIEW_ROWS
        )

    async def test_preview_performs_no_database_writes(self) -> None:
        session = _SkuSession()
        await prepare_catalog_import(
            session,
            BUSINESS_A,
            filename="catalog.csv",
            content=b"name\nWidget\n",
        )
        self.assertEqual(session.add_calls, 0)
        self.assertEqual(session.flush_calls, 0)
        self.assertEqual(session.commit_calls, 0)

    async def test_utf8_bom_and_standard_csv_quoting_work(self) -> None:
        content = '\ufeffname,description\nWidget,"Line one, line two"\n'.encode()
        prepared = await prepare_catalog_import(
            _SkuSession(),
            BUSINESS_A,
            filename="catalog.csv",
            content=content,
        )
        self.assertEqual(prepared.items[0].description, "Line one, line two")


class CatalogImportFileSafetyTests(unittest.TestCase):
    def test_malformed_or_undecodable_csv_is_rejected(self) -> None:
        for content in (b'name\n"unterminated\n', b"name\n\xff\n"):
            with (
                self.subTest(content=content),
                self.assertRaises(CatalogImportFileError),
            ):
                parse_catalog_import_file("catalog.csv", content)

    def test_invalid_xlsx_is_rejected_even_with_correct_extension(self) -> None:
        with self.assertRaises(CatalogImportFileError):
            parse_catalog_import_file("catalog.xlsx", b"not an xlsx")

    def test_unsupported_extensions_are_rejected(self) -> None:
        for filename in ("catalog.xls", "catalog.zip", "catalog.xlsm", "catalog.bin"):
            with (
                self.subTest(filename=filename),
                self.assertRaises(CatalogImportFileError),
            ):
                parse_catalog_import_file(filename, b"name\nWidget\n")

    def test_maximum_upload_size_is_enforced_before_parsing(self) -> None:
        with self.assertRaises(CatalogImportTooLargeError):
            parse_catalog_import_file(
                "catalog.csv",
                b"x" * (MAX_CATALOG_IMPORT_BYTES + 1),
            )

    def test_maximum_row_count_is_enforced(self) -> None:
        rows = ["name"] + [f"Item {index}" for index in range(2001)]
        with self.assertRaisesRegex(CatalogImportFileError, "at most 2000"):
            parse_catalog_import_file("catalog.csv", ("\n".join(rows) + "\n").encode())

    def test_missing_header_name_empty_file_and_header_only_are_rejected(self) -> None:
        cases = (
            ("catalog.csv", b"price\n1.00\n"),
            ("catalog.csv", b""),
            ("catalog.csv", b"name\n"),
        )
        for filename, content in cases:
            with (
                self.subTest(content=content),
                self.assertRaises(CatalogImportFileError),
            ):
                parse_catalog_import_file(filename, content)

    def test_ambiguous_alias_columns_are_rejected(self) -> None:
        with self.assertRaisesRegex(CatalogImportFileError, "Multiple columns"):
            parse_catalog_import_file(
                "catalog.csv",
                b"name,product name\nOne,Two\n",
            )

    def test_xlsx_is_loaded_read_only_and_data_only(self) -> None:
        workbook = Workbook()
        workbook.active.append(["name", "price"])
        workbook.active.append(["Widget", "=1+1"])
        content = _workbook_bytes(workbook)

        from app.services import catalog_import

        real_load_workbook = catalog_import.load_workbook
        calls: list[dict[str, object]] = []

        def recording_load_workbook(*args, **kwargs):
            calls.append(kwargs)
            return real_load_workbook(*args, **kwargs)

        with patch.object(catalog_import, "load_workbook", recording_load_workbook):
            parsed = parse_catalog_import_file("catalog.xlsx", content)

        self.assertEqual(
            calls,
            [{"read_only": True, "data_only": True, "keep_links": False}],
        )
        self.assertEqual(parsed.rows[0].values["name"], "Widget")
        self.assertIsNone(parsed.rows[0].values["price"])

    def test_file_metadata_never_contains_raw_content(self) -> None:
        parsed = parse_catalog_import_file(
            "../../catalog.csv",
            b"name\nWidget\n",
        )
        dumped = parsed.metadata.model_dump()
        self.assertEqual(dumped["filename"], "catalog.csv")
        self.assertNotIn("content", dumped)
        self.assertNotIn("data", dumped)
        self.assertNotIn("bytes", dumped)


class _ScalarResult:
    def __init__(self, values: list[str]) -> None:
        self._values = values

    def all(self) -> list[str]:
        return self._values


class _SkuSession:
    def __init__(self, existing: dict[UUID, set[str]] | None = None) -> None:
        self.existing = existing or {}
        self.requested_business_id: UUID | None = None
        self.add_calls = 0
        self.flush_calls = 0
        self.commit_calls = 0

    async def scalars(self, statement) -> _ScalarResult:
        params = statement.compile().params
        business_id = next(
            value for value in params.values() if isinstance(value, UUID)
        )
        candidates = next(
            (
                set(value)
                for value in params.values()
                if isinstance(value, (list, tuple, set, frozenset))
            ),
            set(),
        )
        self.requested_business_id = business_id
        return _ScalarResult(sorted(self.existing.get(business_id, set()) & candidates))


def _workbook_bytes(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()
